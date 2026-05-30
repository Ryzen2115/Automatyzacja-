"""
Raport Generator — automatyzacja obsługi plików Excel
=====================================================
Wczytuje dane pracowników i miesięczne dane szczegółowe,
waliduje, scala i generuje raport wynikowy.
"""

import os
import glob
import logging
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ---------------------------------------------------------------------------
# Konfiguracja logowania
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Stałe
# ---------------------------------------------------------------------------
WYMAGANE_KOLUMNY_BAZOWE = {"ID_pracownika", "Imie", "Nazwisko", "Dział", "Stawka_godzinowa", "Prowizja_procent"}
WYMAGANE_KOLUMNY_DANE = {"ID_pracownika", "Miesiac", "Godziny_pracy", "Sprzedaz"}

KOLOR_NAGLOWEK = "1F4E79"      # ciemny niebieski
KOLOR_PODSUMOWANIE = "2E86AB"  # niebieski
KOLOR_PARZYSTE = "D6E4F0"      # jasnoniebieski
KOLOR_BLAD = "FFCCCC"          # jasnoczerony


# ---------------------------------------------------------------------------
# 1. Wczytywanie i walidacja
# ---------------------------------------------------------------------------

def wczytaj_plik_bazowy(sciezka: str) -> pd.DataFrame:
    """Wczytuje plik bazowy z danymi pracowników i sprawdza kolumny."""
    log.info(f"Wczytywanie pliku bazowego: {sciezka}")
    df = pd.read_excel(sciezka, dtype={"ID_pracownika": str})
    brakujace = WYMAGANE_KOLUMNY_BAZOWE - set(df.columns)
    if brakujace:
        raise ValueError(f"Brakujące kolumny w pliku bazowym: {brakujace}")
    log.info(f"  → Wczytano {len(df)} pracowników.")
    return df


def wczytaj_plik_danych(sciezka: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Wczytuje plik z danymi miesięcznymi.
    Zwraca (poprawne_rekordy, błędne_rekordy).
    """
    log.info(f"Wczytywanie pliku danych: {sciezka}")
    df = pd.read_excel(sciezka, dtype={"ID_pracownika": str, "Godziny_pracy": object})

    brakujace = WYMAGANE_KOLUMNY_DANE - set(df.columns)
    if brakujace:
        log.warning(f"  ⚠ Pominięto plik — brakujące kolumny: {brakujace}")
        return pd.DataFrame(), pd.DataFrame()

    # Walidacja: konwersja Godziny_pracy na liczby
    df["Godziny_pracy_num"] = pd.to_numeric(df["Godziny_pracy"], errors="coerce")
    maska_bledow = df["Godziny_pracy_num"].isna()

    bledy = df[maska_bledow].copy()
    bledy["Plik_zrodlowy"] = os.path.basename(sciezka)
    bledy["Powod_bledu"] = "Nieprawidłowa wartość w kolumnie Godziny_pracy"

    poprawne = df[~maska_bledow].copy()
    poprawne["Godziny_pracy"] = poprawne["Godziny_pracy_num"]
    poprawne.drop(columns=["Godziny_pracy_num"], inplace=True)

    if not bledy.empty:
        log.warning(f"  ⚠ Znaleziono {len(bledy)} błędnych rekordów.")
    log.info(f"  → {len(poprawne)} poprawnych rekordów.")
    return poprawne, bledy


def wczytaj_wszystkie_dane(folder: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Wczytuje wszystkie pliki czesc_danych_*.xlsx z folderu."""
    pliki = sorted(glob.glob(os.path.join(folder, "czesc_danych_*.xlsx")))
    if not pliki:
        raise FileNotFoundError(f"Nie znaleziono plików czesc_danych_*.xlsx w: {folder}")

    wszystkie_poprawne = []
    wszystkie_bledy = []
    for sciezka in pliki:
        poprawne, bledy = wczytaj_plik_danych(sciezka)
        if not poprawne.empty:
            wszystkie_poprawne.append(poprawne)
        if not bledy.empty:
            wszystkie_bledy.append(bledy)

    df_dane = pd.concat(wszystkie_poprawne, ignore_index=True) if wszystkie_poprawne else pd.DataFrame()
    df_bledy = pd.concat(wszystkie_bledy, ignore_index=True) if wszystkie_bledy else pd.DataFrame()
    log.info(f"Łącznie wczytano {len(df_dane)} poprawnych rekordów z {len(pliki)} plików.")
    return df_dane, df_bledy


# ---------------------------------------------------------------------------
# 2. Scalanie i obliczenia
# ---------------------------------------------------------------------------

def scal_i_oblicz(df_baza: pd.DataFrame, df_dane: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Scala dane bazowe z danymi miesięcznymi.
    Wylicza wynagrodzenie, prowizje i łączny koszt.
    Zwraca (szczegoly, podsumowanie_na_pracownika).
    """
    # Walidacja ID — oznaczenie rekordów bez odpowiednika w bazie
    znane_id = set(df_baza["ID_pracownika"])
    nieznane = df_dane[~df_dane["ID_pracownika"].isin(znane_id)].copy()
    if not nieznane.empty:
        log.warning(f"  ⚠ {len(nieznane)} rekordów z nieznanym ID_pracownika — zostaną pominięte.")

    df_dane = df_dane[df_dane["ID_pracownika"].isin(znane_id)].copy()

    # Złączenie z bazą
    df = df_dane.merge(df_baza, on="ID_pracownika", how="left")

    # Obliczenia
    df["Wynagrodzenie_podstawowe"] = df["Godziny_pracy"] * df["Stawka_godzinowa"]
    df["Prowizja"] = df["Sprzedaz"] * df["Prowizja_procent"] / 100
    df["Laczny_koszt"] = df["Wynagrodzenie_podstawowe"] + df["Prowizja"]

    # Porządek kolumn szczegółowych
    kolumny = [
        "ID_pracownika", "Imie", "Nazwisko", "Dział", "Stanowisko",
        "Miesiac", "Godziny_pracy", "Stawka_godzinowa",
        "Wynagrodzenie_podstawowe", "Sprzedaz", "Prowizja_procent",
        "Prowizja", "Laczny_koszt",
    ]
    if "Nieobecnosci_dni" in df.columns:
        kolumny.append("Nieobecnosci_dni")
    df_szczegoly = df[kolumny].sort_values(["ID_pracownika", "Miesiac"])

    # Podsumowanie na pracownika
    agg = {
        "Godziny_pracy": "sum",
        "Wynagrodzenie_podstawowe": "sum",
        "Sprzedaz": "sum",
        "Prowizja": "sum",
        "Laczny_koszt": "sum",
    }
    if "Nieobecnosci_dni" in df.columns:
        agg["Nieobecnosci_dni"] = "sum"

    df_podsum = (
        df.groupby(["ID_pracownika", "Imie", "Nazwisko", "Dział", "Stanowisko", "Stawka_godzinowa"])
        .agg(agg)
        .reset_index()
        .sort_values("Laczny_koszt", ascending=False)
    )

    log.info("Obliczenia zakończone.")
    return df_szczegoly, df_podsum


def podsumowanie_dzialowe(df_szczegoly: pd.DataFrame) -> pd.DataFrame:
    """Agreguje dane według działu."""
    agg = {
        "Godziny_pracy": "sum",
        "Wynagrodzenie_podstawowe": "sum",
        "Sprzedaz": "sum",
        "Prowizja": "sum",
        "Laczny_koszt": "sum",
    }
    df = (
        df_szczegoly.groupby("Dział")
        .agg(agg)
        .reset_index()
        .sort_values("Laczny_koszt", ascending=False)
    )
    df.insert(1, "Liczba_pracownikow", df_szczegoly.groupby("Dział")["ID_pracownika"].nunique().values)
    return df


# ---------------------------------------------------------------------------
# 3. Formatowanie i zapis raportu
# ---------------------------------------------------------------------------

def _styl_naglowka(komorka, kolor_hex: str = KOLOR_NAGLOWEK):
    komorka.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    komorka.fill = PatternFill("solid", start_color=kolor_hex)
    komorka.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _ramka_cienka():
    cienka = Side(style="thin", color="AAAAAA")
    return Border(left=cienka, right=cienka, top=cienka, bottom=cienka)


def _format_waluty(arkusz, zakres_komorek):
    for wiersz in arkusz[zakres_komorek]:
        for kom in wiersz:
            kom.number_format = '#,##0.00 "zł"'


def _dopasuj_szerokosc(arkusz, min_szer=10, max_szer=35):
    for kolumna in arkusz.columns:
        max_dl = 0
        for kom in kolumna:
            if kom.value:
                max_dl = max(max_dl, len(str(kom.value)))
        arkusz.column_dimensions[get_column_letter(kolumna[0].column)].width = min(max(max_dl + 2, min_szer), max_szer)


def zapisz_arkusz_danych(wb: openpyxl.Workbook, df: pd.DataFrame, nazwa: str, tytul: str):
    """Zapisuje DataFrame do arkusza z formatowaniem."""
    ws = wb.create_sheet(nazwa)
    ws.title = nazwa

    # Tytuł
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    ws["A1"] = tytul
    ws["A1"].font = Font(bold=True, name="Arial", size=14, color=KOLOR_NAGLOWEK)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    # Nagłówki
    for col_idx, nazwa_kol in enumerate(df.columns, start=1):
        kom = ws.cell(row=2, column=col_idx, value=nazwa_kol.replace("_", " "))
        _styl_naglowka(kom)
    ws.row_dimensions[2].height = 22

    # Dane z naprzemiennym kolorowaniem
    for row_idx, wiersz in enumerate(df.itertuples(index=False), start=3):
        kolor_tla = KOLOR_PARZYSTE if row_idx % 2 == 0 else "FFFFFF"
        for col_idx, wartosc in enumerate(wiersz, start=1):
            kom = ws.cell(row=row_idx, column=col_idx, value=wartosc)
            kom.fill = PatternFill("solid", start_color=kolor_tla)
            kom.border = _ramka_cienka()
            kom.font = Font(name="Arial", size=10)
            kom.alignment = Alignment(vertical="center")

    # Formatowanie walutowe dla kolumn z kwotami
    kol_walutowe = [
        i + 1 for i, n in enumerate(df.columns)
        if any(s in n for s in ["koszt", "wynagrodzenie", "Wynagrodzenie", "Sprzedaz", "Prowizja", "Stawka"])
    ]
    for col_idx in kol_walutowe:
        for row_idx in range(3, len(df) + 3):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00 "zł"'

    # Wiersz sumy (dla kolumn numerycznych)
    wiersz_sumy = len(df) + 3
    ws.cell(row=wiersz_sumy, column=1, value="SUMA").font = Font(bold=True, name="Arial")
    ws.cell(row=wiersz_sumy, column=1).fill = PatternFill("solid", start_color=KOLOR_PODSUMOWANIE)
    ws.cell(row=wiersz_sumy, column=1).font = Font(bold=True, color="FFFFFF", name="Arial")

    for col_idx in range(1, len(df.columns) + 1):
        kom = ws.cell(row=wiersz_sumy, column=col_idx)
        kom.fill = PatternFill("solid", start_color=KOLOR_PODSUMOWANIE)
        kom.border = _ramka_cienka()

    for col_idx, nazwa_kol in enumerate(df.columns, start=1):
        if pd.api.types.is_numeric_dtype(df[nazwa_kol]):
            litera = get_column_letter(col_idx)
            ws.cell(row=wiersz_sumy, column=col_idx).value = f"=SUM({litera}3:{litera}{wiersz_sumy - 1})"
            ws.cell(row=wiersz_sumy, column=col_idx).font = Font(bold=True, color="FFFFFF", name="Arial")
            if col_idx in kol_walutowe:
                ws.cell(row=wiersz_sumy, column=col_idx).number_format = '#,##0.00 "zł"'

    # Zamrożenie nagłówków
    ws.freeze_panes = "A3"
    _dopasuj_szerokosc(ws)


def dodaj_wykres(wb: openpyxl.Workbook, df_dzial: pd.DataFrame):
    """Dodaje wykres słupkowy kosztów według działów."""
    ws_wykres = wb.create_sheet("Wykres")
    ws_wykres.title = "Wykres"

    # Dane pomocnicze do wykresu
    ws_wykres["A1"] = "Dział"
    ws_wykres["B1"] = "Łączny koszt (zł)"
    for i, (_, row) in enumerate(df_dzial.iterrows(), start=2):
        ws_wykres.cell(row=i, column=1, value=row["Dział"])
        ws_wykres.cell(row=i, column=2, value=row["Laczny_koszt"])

    chart = BarChart()
    chart.type = "col"
    chart.title = "Koszty wynagrodzeń według działów"
    chart.y_axis.title = "Kwota (zł)"
    chart.x_axis.title = "Dział"
    chart.style = 10
    chart.width = 18
    chart.height = 12

    n = len(df_dzial) + 1
    data_ref = Reference(ws_wykres, min_col=2, min_row=1, max_row=n)
    cats_ref = Reference(ws_wykres, min_col=1, min_row=2, max_row=n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws_wykres.add_chart(chart, "D2")
    ws_wykres.sheet_state = "visible"


def zapisz_arkusz_bledow(wb: openpyxl.Workbook, df_bledy: pd.DataFrame):
    """Zapisuje błędne rekordy do osobnego arkusza."""
    ws = wb.create_sheet("Błędy walidacji")
    ws["A1"] = "Błędy walidacji danych"
    ws["A1"].font = Font(bold=True, color="CC0000", name="Arial", size=13)
    ws.merge_cells(f"A1:{get_column_letter(len(df_bledy.columns))}1")

    for col_idx, nazwa in enumerate(df_bledy.columns, start=1):
        kom = ws.cell(row=2, column=col_idx, value=nazwa.replace("_", " "))
        _styl_naglowka(kom, "CC0000")

    for row_idx, wiersz in enumerate(df_bledy.itertuples(index=False), start=3):
        for col_idx, wartosc in enumerate(wiersz, start=1):
            kom = ws.cell(row=row_idx, column=col_idx, value=wartosc)
            kom.fill = PatternFill("solid", start_color=KOLOR_BLAD)
            kom.border = _ramka_cienka()
            kom.font = Font(name="Arial", size=10)

    _dopasuj_szerokosc(ws)


def zapisz_raport(
    sciezka_wyjsciowa: str,
    df_szczegoly: pd.DataFrame,
    df_podsum_pracownik: pd.DataFrame,
    df_podsum_dzial: pd.DataFrame,
    df_bledy: pd.DataFrame,
):
    """Generuje plik Excel raportu z wieloma arkuszami."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # usuń domyślny pusty arkusz

    # Arkusz tytułowy
    ws_info = wb.create_sheet("Informacje")
    ws_info["A1"] = "RAPORT WYNAGRODZEŃ PRACOWNIKÓW"
    ws_info["A1"].font = Font(bold=True, name="Arial", size=16, color=KOLOR_NAGLOWEK)
    ws_info["A3"] = f"Data wygenerowania:"
    ws_info["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_info["A4"] = "Liczba pracowników:"
    ws_info["B4"] = df_podsum_pracownik["ID_pracownika"].nunique()
    ws_info["A5"] = "Liczba rekordów (miesięcy):"
    ws_info["B5"] = len(df_szczegoly)
    ws_info["A6"] = "Łączny koszt wynagrodzeń:"
    ws_info["B6"] = f'=SUMA'  # placeholder — wypełnimy wartością
    total = df_podsum_pracownik["Laczny_koszt"].sum()
    ws_info["B6"] = total
    ws_info["B6"].number_format = '#,##0.00 "zł"'
    for r in range(3, 7):
        ws_info.cell(row=r, column=1).font = Font(bold=True, name="Arial")
        ws_info.cell(row=r, column=2).font = Font(name="Arial")
    ws_info.column_dimensions["A"].width = 28
    ws_info.column_dimensions["B"].width = 22

    # Arkusze danych
    zapisz_arkusz_danych(wb, df_szczegoly, "Dane szczegółowe", "Szczegółowe dane miesięczne pracowników")
    zapisz_arkusz_danych(wb, df_podsum_pracownik, "Podsumowanie pracownicy", "Podsumowanie wynagrodzeń — pracownicy")
    zapisz_arkusz_danych(wb, df_podsum_dzial, "Podsumowanie działy", "Podsumowanie kosztów według działów")
    dodaj_wykres(wb, df_podsum_dzial)

    if not df_bledy.empty:
        zapisz_arkusz_bledow(wb, df_bledy)

    wb.save(sciezka_wyjsciowa)
    log.info(f"Raport zapisany: {sciezka_wyjsciowa}")


# ---------------------------------------------------------------------------
# 4. Główna funkcja
# ---------------------------------------------------------------------------

def generuj_raport(
    plik_bazowy: str,
    folder_danych: str,
    plik_wyjsciowy: str,
):
    """Pełny pipeline: wczytanie → walidacja → scalenie → obliczenia → raport."""
    log.info("=" * 60)
    log.info("START GENEROWANIA RAPORTU")
    log.info("=" * 60)

    df_baza = wczytaj_plik_bazowy(plik_bazowy)
    df_dane, df_bledy = wczytaj_wszystkie_dane(folder_danych)

    if df_dane.empty:
        log.error("Brak poprawnych danych do przetworzenia. Przerywam.")
        return

    df_szczegoly, df_podsum_pracownik = scal_i_oblicz(df_baza, df_dane)
    df_podsum_dzial = podsumowanie_dzialowe(df_szczegoly)

    os.makedirs(os.path.dirname(plik_wyjsciowy) or ".", exist_ok=True)
    zapisz_raport(plik_wyjsciowy, df_szczegoly, df_podsum_pracownik, df_podsum_dzial, df_bledy)

    log.info("=" * 60)
    log.info("RAPORT WYGENEROWANY POMYŚLNIE")
    log.info("=" * 60)


# ---------------------------------------------------------------------------
# Uruchomienie
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    generuj_raport(
        plik_bazowy="data/dane_glowne.xlsx",
        folder_danych="data/",
        plik_wyjsciowy="output/raport_koncowy.xlsx",
    )
